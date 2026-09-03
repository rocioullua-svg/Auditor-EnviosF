import streamlit as st
import pandas as pd
import numpy as np
import io
import openpyxl
import uuid
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración de la página
st.set_page_config(page_title="Auditoría Flexit v4.1", page_icon="📊", layout="wide")

st.title("📊 Auditoría Definitiva de Envíos v4.1")
st.markdown("Verifica cobros, busca exactamente **FLEX IT**, incluye vacíos y utiliza búsqueda inteligente en cascada.")

# Tarifas Oficiales Flexit
TARIFA_CABA = 4610.99
TARIFA_GBA1 = 7370.99
TARIFA_GBA2 = 10245.99
TARIFAS_VALIDAS = [TARIFA_CABA, TARIFA_GBA1, TARIFA_GBA2]

def cargar_archivo(file_uploader):
    """Escáner inteligente de archivos"""
    nombre_archivo = file_uploader.name
    archivo_bytes = file_uploader.read()
    
    if nombre_archivo.lower().endswith(('.xls', '.xlsx')):
        df = pd.read_excel(io.BytesIO(archivo_bytes), header=None)
    else:
        df = pd.read_csv(io.BytesIO(archivo_bytes), encoding='latin-1', header=None, on_bad_lines='skip')
        
    header_row_idx = None
    for i, row in df.head(15).iterrows():
        if row.astype(str).str.contains('Tracking|Pedido|venta ML|ID venta|Cliente', case=False, na=False).any():
            header_row_idx = i
            break
            
    if header_row_idx is not None:
        df.columns = df.iloc[header_row_idx].fillna('Columna_Sin_Nombre')
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        
    df.columns = df.columns.astype(str).str.strip()
    return df

def limpiar_texto(columna):
    res = columna.astype(str).str.strip().str.replace('.0', '', regex=False).str.upper()
    # Asigna un ID único a los vacíos para que no se crucen entre sí por error
    return res.apply(lambda x: str(uuid.uuid4()) if x in ['NAN', 'NONE', 'NULL', ''] else x)

def limpiar_nombre(columna):
    res = columna.astype(str).str.upper().str.replace(r'[^A-Z0-9\s]', '', regex=True).str.replace(r'\s+', ' ', regex=True).str.strip()
    return res.apply(lambda x: str(uuid.uuid4()) if x in ['NAN', 'NONE', 'NULL', ''] else x)

# --- INTERFAZ DE CARGA ---
col1, col2 = st.columns(2)
with col1:
    archivo_prov = st.file_uploader("📁 1. Archivo Proveedor (Flexit)", type=['csv', 'xls', 'xlsx'])
with col2:
    archivo_int = st.file_uploader("📁 2. Archivo Interno (Sistema)", type=['csv', 'xls', 'xlsx'])

# --- PROCESAMIENTO ---
if archivo_prov and archivo_int:
    with st.spinner("⚙️ Ejecutando Motor de Búsqueda y auditando..."):
        
        df_prov = cargar_archivo(archivo_prov)
        df_int = cargar_archivo(archivo_int)
        
        # Filtro Inteligente de Transporte: Busca exactamente FLEX IT o celdas vacías
        if 'Pedido - Transportista' in df_int.columns:
            transporte = df_int['Pedido - Transportista'].astype(str).str.upper()
            condicion_flexit = transporte.str.contains('FLEX IT')
            condicion_vacio = df_int['Pedido - Transportista'].isna() | (transporte == 'NAN') | (transporte == '')
            
            df_int_flexit = df_int[condicion_flexit | condicion_vacio].copy()
        else:
            df_int_flexit = df_int.copy()

        # Generación de Llaves de Cruce Seguras (Proveedor)
        df_prov['T_P'] = limpiar_texto(df_prov.get('Número Tracking', pd.Series(dtype=str)))
        df_prov['O_P'] = limpiar_texto(df_prov.get('ID venta ML', pd.Series(dtype=str)))
        
        col_nombre_prov = next((col for col in df_prov.columns if 'destinatario' in col.lower() or 'cliente' in col.lower()), 'Nombre Destinatario')
        df_prov['N_P'] = limpiar_nombre(df_prov.get(col_nombre_prov, pd.Series(dtype=str)))
        if 'CP' in df_prov.columns:
            df_prov['CP_Num'] = pd.to_numeric(df_prov['CP'].astype(str).str.extract(r'(\d+)')[0], errors='coerce').fillna(0)

        # Generación de Llaves de Cruce Seguras (Sistema Interno)
        df_int_flexit['T_I'] = limpiar_texto(df_int_flexit.get('Tracking Code', pd.Series(dtype=str)))
        df_int_flexit['O_I'] = limpiar_texto(df_int_flexit.get('Nro Pedido', pd.Series(dtype=str)))
        
        col_nombre_int = next((col for col in df_int_flexit.columns if 'cliente' in col.lower() or 'destinatario' in col.lower()), 'Cliente')
        df_int_flexit['N_I'] = limpiar_nombre(df_int_flexit.get(col_nombre_int, pd.Series(dtype=str)))

        # === MOTOR DE BÚSQUEDA EN CASCADA (5 NIVELES) ===
        matches = []
        
        # Nivel 1: Tracking (Prov) == Tracking (Int)
        m1 = pd.merge(df_prov, df_int_flexit, left_on='T_P', right_on='T_I', how='inner')
        m1['Nivel_Cruce'] = 'Nivel 1 (Trackings Exactos)'
        matches.append(m1)
        
        un_p = df_prov[~df_prov['T_P'].isin(m1['T_P'])]
        un_i = df_int_flexit[~df_int_flexit['T_I'].isin(m1['T_I'])]

        # Nivel 2: ID Venta (Prov) == Nro Pedido (Int)
        m2 = pd.merge(un_p, un_i, left_on='O_P', right_on='O_I', how='inner')
        m2['Nivel_Cruce'] = 'Nivel 2 (ID Venta Exacto)'
        matches.append(m2)
        
        un_p = un_p[~un_p['O_P'].isin(m2['O_P'])]
        un_i = un_i[~un_i['O_I'].isin(m2['O_I'])]

        # Nivel 3: Tracking (Prov) == Nro Pedido (Int) -> Cruzado
        m3 = pd.merge(un_p, un_i, left_on='T_P', right_on='O_I', how='inner')
        m3['Nivel_Cruce'] = 'Nivel 3 (Tracking Prov -> Pedido Int)'
        matches.append(m3)
        
        un_p = un_p[~un_p['T_P'].isin(m3['T_P'])]
        un_i = un_i[~un_i['O_I'].isin(m3['O_I'])]

        # Nivel 4: ID Venta (Prov) == Tracking (Int) -> Cruzado
        m4 = pd.merge(un_p, un_i, left_on='O_P', right_on='T_I', how='inner')
        m4['Nivel_Cruce'] = 'Nivel 4 (ID Venta Prov -> Tracking Int)'
        matches.append(m4)
        
        un_p = un_p[~un_p['O_P'].isin(m4['O_P'])]
        un_i = un_i[~un_i['T_I'].isin(m4['T_I'])]

        # Nivel 5: Rescate por Nombre del Cliente
        m5 = pd.merge(un_p, un_i, left_on='N_P', right_on='N_I', how='inner')
        m5['Nivel_Cruce'] = 'Nivel 5 (Rescate por Nombre Cliente)'
        matches.append(m5)
        
        un_p = un_p[~un_p['N_P'].isin(m5['N_P'])]
        un_i = un_i[~un_i['N_I'].isin(m5['N_I'])]

        # Consolidar Resultados
        un_p['_merge'] = 'left_only'
        un_p['Nivel_Cruce'] = 'No Encontrado en Sistema'
        un_i['_merge'] = 'right_only'
        un_i['Nivel_Cruce'] = 'No Facturado por Flexit'

        all_matched = pd.concat(matches, ignore_index=True)
        if not all_matched.empty:
            all_matched['_merge'] = 'both'

        cruce = pd.concat([all_matched, un_p, un_i], ignore_index=True)

        # Configuración Monetaria
        col_precio_prov = 'Precio Facturado' if 'Precio Facturado' in cruce.columns else 'Precio'
        for col in [col_precio_prov, 'Costo de Envío', 'Costo de Envío Cliente']:
            if col in cruce.columns:
                cruce[col] = pd.to_numeric(cruce[col].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
            else:
                cruce[col] = 0

        # Lógica 1: Auditoría a Flexit (Proveedor)
        def auditar_flexit(row):
            cobro = row[col_precio_prov]
            cp = row.get('CP_Num', 0)
            
            if row['_merge'] == 'left_only':
                return 'Fantasma (Facturado pero No en Sistema)'
            elif row['_merge'] == 'right_only':
                return 'Omitido (No facturado por Flexit)'
                
            es_caba = 1000 <= cp <= 1499
            es_tarifa_oficial = any(abs(cobro - tarifa) <= 10 for tarifa in TARIFAS_VALIDAS)
            
            if es_caba:
                if (cobro - TARIFA_CABA) > 50:
                    return 'Reclamo: Sobreprecio CABA'
                return 'Cobro OK (CABA)'
            else:
                if es_tarifa_oficial:
                    return 'Cobro OK (Tarifa Oficial)'
                else:
                    return 'Reclamo: Tarifa Irregular / Inventada'
                    
        cruce['Estado_Flexit'] = cruce.apply(auditar_flexit, axis=1)

        # Lógica 2: Auditoría del Sistema Interno (Zonas Faltantes)
        def auditar_sistema(row):
            if row['_merge'] == 'right_only':
                return 'N/A (No facturado)'
            if row['Costo de Envío'] == 0:
                return 'Falta Zona en Sistema ($0)'
            return 'Cotizado OK'
            
        cruce['Alerta_Sistema_Interno'] = cruce.apply(auditar_sistema, axis=1)

        # Lógica 3: Cálculo de Reclamos
        def calcular_reclamo(row):
            estado = row['Estado_Flexit']
            cobro = row[col_precio_prov]
            costo_sistema = row['Costo de Envío']
            
            if estado == 'Fantasma (Facturado pero No en Sistema)':
                return cobro
            elif estado == 'Reclamo: Sobreprecio CABA':
                return round(cobro - TARIFA_CABA, 2)
            elif estado == 'Reclamo: Tarifa Irregular / Inventada':
                if costo_sistema > 0:
                    return round(cobro - costo_sistema, 2)
                else:
                    return round(max(0, cobro - TARIFA_GBA2), 2)
            return 0.0

        cruce['Monto_a_Reclamar'] = cruce.apply(calcular_reclamo, axis=1)

        # Ordenar columnas para el reporte final
        columnas_deseadas = ['Número Tracking', 'ID venta ML', 'Nro Pedido', col_nombre_int, col_nombre_prov,
                             'Estado_Flexit', 'Alerta_Sistema_Interno', 'Nivel_Cruce', 'Monto_a_Reclamar', 
                             col_precio_prov, 'Costo de Envío', 'Costo de Envío Cliente', 
                             'Fecha Venta', 'Localidad', 'CP', 'Provincia']
        columnas_existentes = [col for col in columnas_deseadas if col in cruce.columns]
        
        reporte_final = cruce[columnas_existentes].copy().fillna('N/A')
        
        # Filtros de Pestañas
        df_reclamos = reporte_final[reporte_final['Monto_a_Reclamar'] > 0].copy()
        
        df_zonas = cruce[cruce['Alerta_Sistema_Interno'] == 'Falta Zona en Sistema ($0)'].copy()
        if not df_zonas.empty and all(col in df_zonas.columns for col in ['Provincia', 'Localidad', 'CP']):
            resumen_zonas = df_zonas.groupby(['Provincia', 'Localidad', 'CP']).size().reset_index(name='Veces_No_Cotizado')
            resumen_zonas = resumen_zonas.sort_values(by='Veces_No_Cotizado', ascending=False)
        else:
            resumen_zonas = pd.DataFrame(columns=['Provincia', 'Localidad', 'CP', 'Veces_No_Cotizado'])

        # --- MÉTRICAS EN PANTALLA ---
        total_prov = cruce[cruce['_merge'] != 'right_only'][col_precio_prov].sum()
        total_reclamos = df_reclamos['Monto_a_Reclamar'].sum() if not df_reclamos.empty else 0
        viajes_sin_zona = len(df_zonas)
        viajes_rescatados = len(cruce[cruce['Nivel_Cruce'].str.contains('Nivel 3|Nivel 4|Nivel 5', na=False)])
        
        st.success("¡Auditoría completada exitosamente!")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Facturado (Flexit)", f"$ {total_prov:,.2f}")
        m2.metric("Total a Reclamar", f"$ {total_reclamos:,.2f}", delta="Fantasmas y Sobreprecios", delta_color="inverse")
        m3.metric("Fallas del Sistema (Costo $0)", f"{viajes_sin_zona} envíos")
        m4.metric("Viajes Rescatados (Cruzados)", f"{viajes_rescatados} envíos")
        
        # --- GENERACIÓN DEL EXCEL ---
        wb = openpyxl.Workbook()
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_data = wb.create_sheet(title="Auditoria General")
        ws_reclamos = wb.create_sheet(title="Reclamos a Flexit") 
        ws_zonas = wb.create_sheet(title="Zonas Faltantes Internas")     
        
        header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
        
        def dar_formato_tabla(ws, dataframe):
            if dataframe.empty:
                ws.append(["No hay registros para mostrar."])
                return
            for r in dataframe_to_rows(dataframe, index=False, header=True):
                ws.append(r)
            for col in ws.iter_cols(min_row=1, max_row=1, max_col=ws.max_column):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    col_name = str(ws.cell(row=1, column=cell.column).value)
                    if any(moneda in col_name for moneda in ['Precio', 'Costo', 'Monto']):
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
            for i, col in enumerate(dataframe.columns):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 22

        dar_formato_tabla(ws_data, reporte_final)
        dar_formato_tabla(ws_reclamos, df_reclamos)
        dar_formato_tabla(ws_zonas, resumen_zonas)
        
        ws_dash.sheet_view.showGridLines = False
        ws_dash['B2'] = "Dashboard de Control - Flexit"
        ws_dash['B2'].font = Font(size=16, bold=True, color="2F4F4F")
        ws_dash['B4'] = "1. Total Facturado por Proveedor:"
        ws_dash['C4'] = total_prov
        ws_dash['B5'] = "2. Dinero Total a RECLAMAR:"
        ws_dash['C5'] = total_reclamos
        ws_dash['C5'].font = Font(color="B22222", bold=True)
        ws_dash['B6'] = "3. Viajes con Zona sin cargar ($0):"
        ws_dash['C6'] = viajes_sin_zona
        ws_dash['B7'] = "4. Viajes rescatados por Nombres/Códigos:"
        ws_dash['C7'] = viajes_rescatados
        
        for r in range(4, 8):
            ws_dash[f'B{r}'].font = Font(bold=True)
            if r != 6 and r != 7: ws_dash[f'C{r}'].number_format = '$#,##0.00'
            
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        st.subheader("📥 Descargar Reporte Final")
        st.download_button(
            label="Descargar Auditoría en Excel",
            data=excel_buffer,
            file_name="Auditoria_Flexit_v4.1.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
